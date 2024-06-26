import streamlit as st
from openpyxl import load_workbook, Workbook
import os
from main import main_page

# Initialize session state for login
if 'logged_in' not in st.session_state:
    st.session_state["logged_in"] = False

if 'username' not in st.session_state:
    st.session_state["username"] = None

# Path to the Excel file
EXCEL_FILE = "users.xlsx"

# Create an Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["username", "password"])  # Add header
    wb.save(EXCEL_FILE)

# Function to read users from the Excel file
def read_users():
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        users.append({"username": row[0], "password": row[1]})
    return users

# Function to add a new user to the Excel file
def add_user(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    ws.append([username, password])
    wb.save(EXCEL_FILE)

# Sign up form
def sign_up():
    st.title("Sign Up")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Sign Up"):
        if username and password:
            users = read_users()
            if any(user["username"] == username for user in users):
                st.error("Username already exists.")
            else:
                add_user(username, password)
                st.success("User registered successfully!")
        else:
            st.error("Please provide a username and password.")

# Login form
def login():
    st.title("Login")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        if username and password:
            users = read_users()
            if any(user["username"] == username and user["password"] == password for user in users):
                st.session_state["logged_in"] = True
                st.session_state["username"] = username
                st.success("Login successful!")
                st.experimental_rerun()  # Refresh the page after successful login
            else:
                st.error("Incorrect username or password.")
        else:
            st.error("Please provide a username and password.")

# Logout function
def logout():
    st.session_state["logged_in"] = False
    st.session_state["username"] = None
    st.success("Logged out successfully!")
    st.experimental_rerun()

# Main application
def main():
    if st.session_state['logged_in']:
        st.sidebar.title(f"Welcome, {st.session_state['username']}!")
        if st.sidebar.button("Logout"):
            logout()
        main_page()
    else:
        st.sidebar.title("Navigation")
        option = st.sidebar.selectbox("Choose a page", ["Sign Up", "Login"])

        if option == "Sign Up":
            sign_up()
        elif option == "Login":
            login()

if __name__ == "__main__":
    main()
